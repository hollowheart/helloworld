#!/usr/bin/env python3
"""
J-Ant .xlsm helpers: Healthy/Late, End FB vs Target FB, empty End FB, plan deviation
(empty End FB or End FB > Target FB), FOT (-Z), delayed (End FB > RC FB; Status not Done/Obsolete),
End FB with open or committed FB Committed Status, no-RFC rows, with Jira Key URLs.
Requires: pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl


def norm(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def norm_raw(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def find_header_row(ws, max_scan: int = 50) -> tuple[int | None, list]:
    for r in range(1, max_scan + 1):
        row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
        cells = [norm(c) for c in row]
        if any("end fb" in c for c in cells if c):
            return r, row
    return None, []


def header_labels(header_row: list) -> list[str]:
    out: list[str] = []
    for i, h in enumerate(header_row):
        if h is None or str(h).strip() == "":
            out.append(f"Column_{i + 1}")
        else:
            out.append(str(h).strip())
    return out


def col_idx(headers: list[str], *parts: str) -> int | None:
    for i, h in enumerate(headers):
        n = norm(h)
        if all(p in n for p in parts):
            return i
    return None


def num(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("n/a", "na", "-", "missing", "none"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_empty_end(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    if not s:
        return True
    if s.lower() in ("n/a", "na", "-"):
        return True
    return False


def is_empty_contact(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    if not s:
        return True
    if s.lower() in ("n/a", "na", "-", "none", "tbd"):
        return True
    return False


def is_empty_or_not_committed_rel_release_status(v: object) -> bool:
    """True if Rel Committed Status is blank/placeholder or explicitly not committed."""
    if v is None:
        return True
    s = str(v).strip()
    if not s:
        return True
    sl = s.lower()
    if sl in ("n/a", "na", "-", "none", "missing"):
        return True
    if "not committed" in norm(s):
        return True
    return False


def is_empty_or_not_committed_fb_status(v) -> bool:
    """True if FB Committed Status is blank/placeholder or explicitly not committed."""
    if v is None:
        return True
    s = str(v).strip()
    if not s:
        return True
    sl = s.lower()
    if sl in ("n/a", "na", "-", "none", "missing"):
        return True
    if "not committed" in norm(s):
        return True
    return False


def key_and_url(ws, excel_row: int, col_key_idx: int) -> tuple[object, str | None]:
    """0-based col_key_idx -> openpyxl column is col_key_idx + 1."""
    c = ws.cell(row=excel_row, column=col_key_idx + 1)
    url = None
    if c.hyperlink is not None:
        url = getattr(c.hyperlink, "target", None)
    return c.value, url


def open_wb(path: Path, copy_to: Path | None):
    if copy_to:
        shutil.copy2(path, copy_to)
        path = copy_to
    return openpyxl.load_workbook(path, data_only=True)


def md_escape_cell(v: object) -> str:
    """Single-line markdown table cell: escape pipes, collapse newlines."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"\s+", " ", s.replace("\n", " "))
    return s.replace("|", "\\|")


def format_key_markdown(rec: dict) -> str:
    k = rec.get("key")
    if k is None or str(k).strip() == "":
        return ""
    ks = md_escape_cell(k)
    u = rec.get("key_url")
    if u:
        return f"[{ks}]({u})"
    return ks


def emit_end_fb_commitment_markdown(
    rows: list[dict],
    *,
    sheet: str,
    end_fb: float,
    feature: str | None,
    committed: bool,
) -> None:
    ef = int(end_fb) if end_fb == int(end_fb) else end_fb
    parts = [f"**Sheet:** {sheet}"]
    if committed:
        parts.append(
            f"**Filter:** End FB = {ef}, FB Committed Status committed (non-open), Type = Competence area"
        )
    else:
        parts.append(
            f"**Filter:** End FB = {ef}, FB Committed Status empty / not committed, Type = Competence area"
        )
    if feature:
        parts.append(f"**Scope:** Summary or Key contains `{feature}`")
    print("\n\n".join(parts))
    print()
    if committed:
        print(
            "| Key | Summary | Competence Area | Assignee | End FB | FB Committed Status | Release committed FB (RC FB) | Risk Status | Stretch Goal Reason |"
        )
        print("| --- | --- | --- | --- | ---: | --- | ---: | --- | --- |")
        for rec in rows:
            print(
                "| "
                + format_key_markdown(rec)
                + " | "
                + md_escape_cell(rec.get("summary"))
                + " | "
                + md_escape_cell(rec.get("competence_area"))
                + " | "
                + md_escape_cell(rec.get("assignee"))
                + " | "
                + md_escape_cell(rec.get("end_fb"))
                + " | "
                + md_escape_cell(rec.get("fb_committed_status"))
                + " | "
                + md_escape_cell(rec.get("release_committed_fb"))
                + " | "
                + md_escape_cell(rec.get("risk_status"))
                + " | "
                + md_escape_cell(rec.get("stretch_goal_reason"))
                + " |"
            )
    else:
        print(
            "| Key | Summary | Competence Area | Assignee | End FB | Release committed FB (RC FB) | Risk Status | Stretch Goal Reason |"
        )
        print("| --- | --- | --- | --- | ---: | ---: | --- | --- |")
        for rec in rows:
            print(
                "| "
                + format_key_markdown(rec)
                + " | "
                + md_escape_cell(rec.get("summary"))
                + " | "
                + md_escape_cell(rec.get("competence_area"))
                + " | "
                + md_escape_cell(rec.get("assignee"))
                + " | "
                + md_escape_cell(rec.get("end_fb"))
                + " | "
                + md_escape_cell(rec.get("release_committed_fb"))
                + " | "
                + md_escape_cell(rec.get("risk_status"))
                + " | "
                + md_escape_cell(rec.get("stretch_goal_reason"))
                + " |"
            )


def emit_no_rfc_end_before_target_markdown(
    rows: list[dict],
    *,
    sheet: str,
    feature: str,
) -> None:
    parts = [
        f"**Sheet:** {sheet}",
        "**Filter:** No RFC - Status not Done/Obsolete, End FB set and numeric, End FB <= Target FB, **Rel Committed Status** empty or not committed; Type = Competence area",
        f"**Scope:** Summary or Key contains `{feature}`",
    ]
    print("\n\n".join(parts))
    print()
    print(
        "| Key | Summary | Status | Competence Area | Assignee | End FB | Target FB | Release committed status |"
    )
    print("| --- | --- | --- | --- | --- | ---: | ---: | --- |")
    for rec in rows:
        print(
            "| "
            + format_key_markdown(rec)
            + " | "
            + md_escape_cell(rec.get("summary"))
            + " | "
            + md_escape_cell(rec.get("status"))
            + " | "
            + md_escape_cell(rec.get("competence_area"))
            + " | "
            + md_escape_cell(rec.get("assignee"))
            + " | "
            + md_escape_cell(rec.get("end_fb"))
            + " | "
            + md_escape_cell(rec.get("target_fb"))
            + " | "
            + md_escape_cell(rec.get("rel_committed_status"))
            + " |"
        )


def emit_plan_deviation_markdown(
    rows: list[dict],
    *,
    sheet: str,
    feature: str | None,
) -> None:
    parts = [
        f"**Sheet:** {sheet}",
        "**Filter:** Plan deviation - **End FB** empty (blank / placeholder) **or** numeric **End FB** > numeric **Target FB**; Type = Competence area; non-empty **Key**; **Status** not Done / Obsolete",
    ]
    if feature:
        parts.append(f"**Scope:** Summary or Key contains `{feature}`")
    print("\n\n".join(parts))
    print()
    print(
        "| Key | Summary | Status | Competence Area | Assignee | Start FB | End FB | Target FB | Risk Status | Release committed status |"
    )
    print("| --- | --- | --- | --- | --- | ---: | ---: | ---: | --- | --- |")
    for rec in rows:
        print(
            "| "
            + format_key_markdown(rec)
            + " | "
            + md_escape_cell(rec.get("summary"))
            + " | "
            + md_escape_cell(rec.get("status"))
            + " | "
            + md_escape_cell(rec.get("competence_area"))
            + " | "
            + md_escape_cell(rec.get("assignee"))
            + " | "
            + md_escape_cell(rec.get("start_fb"))
            + " | "
            + md_escape_cell(rec.get("end_fb"))
            + " | "
            + md_escape_cell(rec.get("target_fb"))
            + " | "
            + md_escape_cell(rec.get("risk_status"))
            + " | "
            + md_escape_cell(rec.get("rel_committed_status"))
            + " |"
        )


def emit_delayed_rc_fb_markdown(
    rows: list[dict],
    *,
    sheet: str,
    feature: str | None,
) -> None:
    parts = [
        f"**Sheet:** {sheet}",
        "**Filter:** Delayed - numeric **End FB** > numeric **RC FB** (Release committed FB); Type = Competence area; **Status** not Done / Obsolete",
    ]
    if feature:
        parts.append(f"**Scope:** Summary or Key contains `{feature}`")
    else:
        parts.append("**Scope:** All competence-area rows on sheet (no `--feature` filter)")
    print("\n\n".join(parts))
    print()
    print(
        "| Key | Summary | Status | Competence Area | Assignee | Release committed FB (RC FB) | End FB | Delay Explanation |"
    )
    print("| --- | --- | --- | --- | --- | ---: | ---: | --- |")
    for rec in rows:
        print(
            "| "
            + format_key_markdown(rec)
            + " | "
            + md_escape_cell(rec.get("summary"))
            + " | "
            + md_escape_cell(rec.get("status"))
            + " | "
            + md_escape_cell(rec.get("competence_area"))
            + " | "
            + md_escape_cell(rec.get("assignee"))
            + " | "
            + md_escape_cell(rec.get("release_committed_fb"))
            + " | "
            + md_escape_cell(rec.get("end_fb"))
            + " | "
            + md_escape_cell(rec.get("delay_explanation"))
            + " |"
        )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--sheet", default="Gantt Chart")
    ap.add_argument("--copy-to", type=Path, default=None)
    ap.add_argument(
        "--filter",
        choices=(
            "healthy-late",
            "healthy-late-exact",
            "end-after-target",
            "empty-end-fb",
            "plan-deviation",
            "delayed-rc-fb",
            "end-fb-not-committed",
            "end-fb-committed",
            "no-rfc-end-before-target",
            "fot",
        ),
        required=True,
    )
    ap.add_argument(
        "--feature",
        type=str,
        default=None,
        help="Substring in Summary (fot) or Summary/Key (delayed-rc-fb, end-fb-*, no-rfc-end-before-target, plan-deviation). Required for fot and no-rfc-end-before-target; optional for delayed-rc-fb (omit for all competence-area rows), plan-deviation, end-fb-not-committed, end-fb-committed.",
    )
    ap.add_argument(
        "--fot-contact",
        choices=("provided", "missing", "both"),
        default="both",
        help="For --filter fot: Contact Person non-empty / empty / both -Z competence rows",
    )
    ap.add_argument(
        "--exclude-ran-sysspec",
        action="store_true",
        help="For --filter fot: drop rows where Competence Area or Summary contains RAN SysSpec",
    )
    ap.add_argument(
        "--end-fb",
        type=float,
        default=None,
        help="For --filter end-fb-not-committed and end-fb-committed: required End FB value (YYWW, e.g. 2611)",
    )
    ap.add_argument(
        "--format",
        choices=("json", "markdown"),
        default="json",
        help="json (default) or markdown table (delayed-rc-fb, end-fb-not-committed, end-fb-committed, no-rfc-end-before-target, plan-deviation)",
    )
    args = ap.parse_args()

    if args.filter in ("end-fb-not-committed", "end-fb-committed") and args.end_fb is None:
        sys.stderr.write(
            f"--end-fb is required for --filter {args.filter} (YYWW, e.g. 2611).\n"
        )
        return 1

    try:
        wb = open_wb(args.workbook, args.copy_to)
    except PermissionError:
        sys.stderr.write("Permission denied; retry with --copy-to a local path.\n")
        return 1

    if args.sheet not in wb.sheetnames:
        sys.stderr.write(f"Unknown sheet {args.sheet!r}. Available: {wb.sheetnames}\n")
        wb.close()
        return 1

    ws = wb[args.sheet]
    hr, hdr = find_header_row(ws)
    if hr is None:
        sys.stderr.write("No header row containing End FB.\n")
        wb.close()
        return 1

    headers = header_labels(hdr)
    i_key = col_idx(headers, "key")
    i_type = col_idx(headers, "type")
    i_sum = col_idx(headers, "summary")
    i_comp = col_idx(headers, "competence")
    i_asg = col_idx(headers, "assignee")
    i_contact = col_idx(headers, "contact", "person")
    if i_contact is None:
        i_contact = col_idx(headers, "contact")
    i_healthy = col_idx(headers, "healthy")
    i_end = col_idx(headers, "end", "fb")
    i_tgt = col_idx(headers, "target", "fb")
    i_rc = col_idx(headers, "rc", "fb")
    i_delay = col_idx(headers, "delay", "explanation")
    i_start = col_idx(headers, "start", "fb")
    i_stat = col_idx(headers, "status")
    i_act = col_idx(headers, "activity", "type")
    i_fb_comm = col_idx(headers, "fb", "committed", "status")
    i_stretch = col_idx(headers, "stretch", "goal", "reason")
    i_risk = col_idx(headers, "risk", "status")
    i_rel_committed = col_idx(headers, "rel", "committed", "status")

    rows_out: list[dict] = []

    for rnum, row in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), start=hr + 1):
        lst = list(row)
        if len(lst) < len(headers):
            lst.extend([None] * (len(headers) - len(lst)))

        def g(i: int | None):
            if i is None or i >= len(lst):
                return None
            return lst[i]

        key_val, key_url = (None, None)
        if i_key is not None:
            key_val, key_url = key_and_url(ws, rnum, i_key)

        if args.filter == "healthy-late":
            if i_healthy is None:
                sys.stderr.write("No Healthy column.\n")
                wb.close()
                return 1
            hv = g(i_healthy)
            if hv is None or "late" not in str(hv).lower():
                continue
        elif args.filter == "healthy-late-exact":
            if i_healthy is None:
                sys.stderr.write("No Healthy column.\n")
                wb.close()
                return 1
            if norm(g(i_healthy)) != "late":
                continue
        elif args.filter == "end-after-target":
            if i_end is None or i_tgt is None:
                sys.stderr.write("Need End FB and Target FB columns.\n")
                wb.close()
                return 1
            end = num(g(i_end))
            tgt = num(g(i_tgt))
            if end is None or tgt is None or end <= tgt:
                continue
        elif args.filter == "empty-end-fb":
            if i_end is None or i_key is None:
                sys.stderr.write("Need End FB and Key columns.\n")
                wb.close()
                return 1
            if not is_empty_end(g(i_end)):
                continue
            if not g(i_key):
                continue
        elif args.filter == "plan-deviation":
            if None in (i_type, i_end, i_tgt, i_key, i_stat):
                sys.stderr.write(
                    "plan-deviation needs Type, End FB, Target FB, Key, and Status columns.\n"
                )
                wb.close()
                return 1
            if norm(g(i_type)) != "competence area":
                continue
            if not g(i_key):
                continue
            st = g(i_stat)
            if st is not None and str(st).strip():
                sn = norm(st)
                if sn == "done" or sn == "obsolete":
                    continue
            if args.feature:
                summ_s = str(g(i_sum) or "") if i_sum is not None else ""
                key_s = str(key_val or "")
                needle = args.feature.lower()
                if needle not in summ_s.lower() and needle not in key_s.lower():
                    continue
            ev = g(i_end)
            dev = is_empty_end(ev)
            if not dev:
                end = num(ev)
                tgt = num(g(i_tgt))
                dev = end is not None and tgt is not None and end > tgt
            if not dev:
                continue
        elif args.filter == "delayed-rc-fb":
            if None in (i_type, i_sum, i_comp, i_key, i_end, i_rc, i_stat):
                sys.stderr.write(
                    "delayed-rc-fb needs Type, Summary, Competence Area, Key, End FB, RC FB, and Status columns.\n"
                )
                wb.close()
                return 1
            if norm(g(i_type)) != "competence area":
                continue
            if not g(i_key):
                continue
            if args.feature:
                summ_s = str(g(i_sum) or "") if i_sum is not None else ""
                key_s = str(key_val or "")
                needle = args.feature.lower()
                if needle not in summ_s.lower() and needle not in key_s.lower():
                    continue
            st = g(i_stat)
            if st is not None and str(st).strip():
                sn = norm(st)
                if sn == "done" or sn == "obsolete":
                    continue
            end = num(g(i_end))
            rc = num(g(i_rc))
            if end is None or rc is None or end <= rc:
                continue
        elif args.filter == "end-fb-not-committed":
            if None in (i_end, i_key, i_fb_comm, i_type):
                sys.stderr.write(
                    "end-fb-not-committed needs Type, End FB, Key, and FB Committed Status columns.\n"
                )
                wb.close()
                return 1
            if norm(g(i_type)) != "competence area":
                continue
            if not g(i_key):
                continue
            end = num(g(i_end))
            if end is None or end != float(args.end_fb):
                continue
            if not is_empty_or_not_committed_fb_status(g(i_fb_comm)):
                continue
            if args.feature:
                summ_s = str(g(i_sum) or "") if i_sum is not None else ""
                key_s = str(key_val or "")
                needle = args.feature.lower()
                if needle not in summ_s.lower() and needle not in key_s.lower():
                    continue
        elif args.filter == "end-fb-committed":
            if None in (i_end, i_key, i_fb_comm, i_type):
                sys.stderr.write(
                    "end-fb-committed needs Type, End FB, Key, and FB Committed Status columns.\n"
                )
                wb.close()
                return 1
            if norm(g(i_type)) != "competence area":
                continue
            if not g(i_key):
                continue
            end = num(g(i_end))
            if end is None or end != float(args.end_fb):
                continue
            if is_empty_or_not_committed_fb_status(g(i_fb_comm)):
                continue
            if args.feature:
                summ_s = str(g(i_sum) or "") if i_sum is not None else ""
                key_s = str(key_val or "")
                needle = args.feature.lower()
                if needle not in summ_s.lower() and needle not in key_s.lower():
                    continue
        elif args.filter == "no-rfc-end-before-target":
            if not args.feature:
                sys.stderr.write("--feature required for no-rfc-end-before-target filter.\n")
                wb.close()
                return 1
            if None in (i_type, i_sum, i_key, i_end, i_tgt, i_rel_committed, i_stat, i_asg, i_comp):
                sys.stderr.write(
                    "no-rfc-end-before-target needs Type, Summary, Key, End FB, Target FB, "
                    "Rel Committed Status, Status, Assignee, Competence Area.\n"
                )
                wb.close()
                return 1
            if norm(g(i_type)) != "competence area":
                continue
            if not g(i_key):
                continue
            summ_s = str(g(i_sum) or "")
            key_s = str(key_val or "")
            needle = args.feature.lower()
            if needle not in summ_s.lower() and needle not in key_s.lower():
                continue
            st = g(i_stat)
            if st is not None and str(st).strip():
                sn = norm(st)
                if sn == "done" or sn == "obsolete":
                    continue
            if is_empty_end(g(i_end)):
                continue
            end = num(g(i_end))
            tgt = num(g(i_tgt))
            if end is None or tgt is None or end > tgt:
                continue
            if not is_empty_or_not_committed_rel_release_status(g(i_rel_committed)):
                continue
        elif args.filter == "fot":
            if not args.feature:
                sys.stderr.write("--feature required for fot filter.\n")
                wb.close()
                return 1
            if None in (i_type, i_sum, i_contact, i_comp, i_asg, i_key):
                sys.stderr.write("fot filter needs Type, Summary, Contact Person, Competence Area, Assignee, Key.\n")
                wb.close()
                return 1
            if norm(g(i_type)) != "competence area":
                continue
            summ = g(i_sum)
            if not summ:
                continue
            sl = str(summ).lower()
            if args.feature.lower() not in sl:
                continue
            if "-z" not in sl:
                continue
            if args.exclude_ran_sysspec:
                comp = g(i_comp)
                if "ran sysspec" in sl:
                    continue
                if comp and "ran sysspec" in str(comp).lower():
                    continue
            has_c = not is_empty_contact(g(i_contact))
            if args.fot_contact == "provided" and not has_c:
                continue
            if args.fot_contact == "missing" and has_c:
                continue

        rec: dict = {"excel_row": rnum, "key": key_val, "key_url": key_url}
        for label, idx in (
            ("type", i_type),
            ("competence_area", i_comp),
            ("assignee", i_asg),
            ("summary", i_sum),
            ("contact_person", i_contact),
            ("healthy", i_healthy),
            ("start_fb", i_start),
            ("end_fb", i_end),
            ("target_fb", i_tgt),
            ("release_committed_fb", i_rc),
            ("fb_committed_status", i_fb_comm),
            ("rel_committed_status", i_rel_committed),
            ("risk_status", i_risk),
            ("stretch_goal_reason", i_stretch),
            ("delay_explanation", i_delay),
            ("status", i_stat),
            ("activity_type", i_act),
        ):
            if idx is not None:
                rec[label] = g(idx)
        rows_out.append(rec)

    wb.close()
    if args.format == "markdown":
        if args.filter == "delayed-rc-fb":
            emit_delayed_rc_fb_markdown(
                rows_out, sheet=args.sheet, feature=args.feature
            )
            return 0
        if args.filter == "no-rfc-end-before-target":
            if not args.feature:
                sys.stderr.write("--feature required with --format markdown for no-rfc-end-before-target.\n")
                return 1
            emit_no_rfc_end_before_target_markdown(
                rows_out, sheet=args.sheet, feature=args.feature
            )
            return 0
        if args.filter == "plan-deviation":
            emit_plan_deviation_markdown(
                rows_out, sheet=args.sheet, feature=args.feature
            )
            return 0
        if args.filter not in ("end-fb-not-committed", "end-fb-committed"):
            sys.stderr.write(
                "--format markdown is only supported for --filter delayed-rc-fb, "
                "end-fb-not-committed, end-fb-committed, no-rfc-end-before-target, or plan-deviation.\n"
            )
            return 1
        emit_end_fb_commitment_markdown(
            rows_out,
            sheet=args.sheet,
            end_fb=float(args.end_fb),
            feature=args.feature,
            committed=args.filter == "end-fb-committed",
        )
        return 0
    print(json.dumps({"count": len(rows_out), "sheet": args.sheet, "rows": rows_out}, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
