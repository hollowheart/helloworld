"""Export CB015362 sub-feature mapping table (from screenshot) to Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\Users\kewang\OneDrive - Nokia\Projects\Cursor\AnaysisOnJiraData\CB015362_subfeature_mapping.xlsx")

HEADERS = [
    "Sub-Feature ID",
    "Sub-Feature Details",
    "Mapped Feature-Part ID",
    "Mapped Feature-Part Details",
    "Mapped User Story ID(s) (both success and failure user stories)",
    "Mapped User Story Details (both success and failure user stories)",
]

ROWS = [
    [
        "CB015362-A",
        "Increase the L1 UL capacity for ABIP/ASOG/ASOH",
        "1",
        "Feature activation/deactivation. Increase the L1 UL capacity for ABIP/ASOG/ASOH",
        "US1",
        "US1",
    ],
    [
        "CB015362-B",
        "remove 16DI and 8RX restriction for full board deployment ABIQ/ASOG/ASOH",
        "3",
        "remove 16DI and 8RX restriction",
        "US1",
        "US1",
    ],
    [
        "CB015362-K",
        "KPI",
        "4",
        "KPI analysis",
        "US1",
        "US1",
    ],
    [
        "CB015362-W",
        "P&C",
        "5",
        "P&C",
        "US1",
        "US1",
    ],
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"

    header_font = Font(bold=True, color="1F4E79")
    body_font = Font(color="2E75B6")
    wrap = Alignment(wrap_text=True, vertical="top")

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = wrap

    for r, row in enumerate(ROWS, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.alignment = wrap

    widths = (18, 52, 12, 52, 36, 36)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print("Saved:", OUT)


if __name__ == "__main__":
    main()
