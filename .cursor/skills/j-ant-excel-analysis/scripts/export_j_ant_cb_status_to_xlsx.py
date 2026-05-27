#!/usr/bin/env python3
"""
Run J-Ant Gantt filters (delayed, plan-deviation, no-RFC, not-committed @ End FB)
for each --feature scope and write one Excel workbook: one sheet per scope.

Requires: openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet name: max 31 chars, no \\ / * ? : [ ]"""
    s = re.sub(r'[\[\]\\/*?:]', "-", str(name).strip())
    if not s:
        s = "Sheet"
    s = s[:31]
    base = s
    n = 2
    while s.lower() in used:
        suffix = f"_{n}"
        s = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(s.lower())
    return s


def run_analyzer(
    py: Path,
    analyzer: Path,
    workbook: Path,
    copy_to: Path | None,
    filter_name: str,
    feature: str | None,
    end_fb: str | None,
) -> dict:
    cmd = [
        sys.executable,
        str(analyzer),
        "--workbook",
        str(workbook),
        "--filter",
        filter_name,
        "--format",
        "json",
    ]
    if copy_to:
        cmd.extend(["--copy-to", str(copy_to)])
    if feature:
        cmd.extend(["--feature", feature])
    if end_fb is not None:
        cmd.extend(["--end-fb", end_fb])
    p = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
    if p.returncode != 0:
        raise RuntimeError(
            f"Analyzer failed ({filter_name}, feature={feature!r}):\n{p.stderr or p.stdout}"
        )
    raw = (p.stdout or "").strip()
    if not raw:
        return {"count": 0, "sheet": "Gantt Chart", "rows": []}
    return json.loads(raw)


def write_hyperlink_cell(ws, row: int, col: int, key: object, url: str | None) -> None:
    c = ws.cell(row=row, column=col, value=key if key is not None else "")
    if url and key is not None and str(key).strip():
        c.hyperlink = url
        c.font = Font(color="0563C1", underline="single")


def append_table(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[dict],
    field_getters: list,
) -> int:
    """Write title + header + rows; field_getters: list of callable(row)->value or (row)->(value, url) for key col."""
    r = start_row
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12)
    r += 1
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h).font = Font(bold=True)
    r += 1
    for rec in rows:
        for j, getter in enumerate(field_getters, start=1):
            out = getter(rec)
            if isinstance(out, tuple) and len(out) == 2:
                val, url = out
                write_hyperlink_cell(ws, r, j, val, url)
            else:
                ws.cell(row=r, column=j, value=out)
        r += 1
    r += 1
    return r


def autofit_columns(ws, max_col: int, max_width: float = 55.0) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            maxlen = max(maxlen, min(len(str(cell.value)), 120))
        ws.column_dimensions[letter].width = min(max(10, maxlen + 2), max_width)


def build_sheet(ws, feature: str, payloads: dict[str, dict]) -> None:
    """payloads keys: delayed, plan_deviation, no_rfc, not_committed_2611"""

    def g(rec, *keys):
        for k in keys:
            if k in rec and rec[k] is not None:
                v = rec[k]
                if v != "":
                    return v
        return ""

    key_pair = lambda rec: (rec.get("key"), rec.get("key_url") or None)

    r = 1
    ws.cell(row=r, column=1, value=f"Scope (Summary or Key contains): {feature}").font = Font(
        bold=True, size=11
    )
    r += 2

    delayed_rows = payloads["delayed"].get("rows") or []
    r = append_table(
        ws,
        r,
        f"1. Delayed items (End FB > RC FB) — {len(delayed_rows)} row(s)",
        [
            "Key",
            "Summary",
            "Status",
            "Competence Area",
            "Assignee",
            "RC FB",
            "End FB",
            "Delay Explanation",
        ],
        delayed_rows,
        [
            lambda rec: key_pair(rec),
            lambda rec: g(rec, "summary"),
            lambda rec: g(rec, "status"),
            lambda rec: g(rec, "competence_area"),
            lambda rec: g(rec, "assignee"),
            lambda rec: g(rec, "release_committed_fb"),
            lambda rec: g(rec, "end_fb"),
            lambda rec: g(rec, "delay_explanation"),
        ],
    )

    pd_rows = payloads["plan_deviation"].get("rows") or []
    r = append_table(
        ws,
        r,
        f"2. Plan deviation (empty End FB or End FB > Target FB) — {len(pd_rows)} row(s)",
        [
            "Key",
            "Summary",
            "Status",
            "Competence Area",
            "Assignee",
            "Start FB",
            "End FB",
            "Target FB",
            "Risk Status",
            "Rel Committed Status",
        ],
        pd_rows,
        [
            lambda rec: key_pair(rec),
            lambda rec: g(rec, "summary"),
            lambda rec: g(rec, "status"),
            lambda rec: g(rec, "competence_area"),
            lambda rec: g(rec, "assignee"),
            lambda rec: g(rec, "start_fb"),
            lambda rec: g(rec, "end_fb"),
            lambda rec: g(rec, "target_fb"),
            lambda rec: g(rec, "risk_status"),
            lambda rec: g(rec, "rel_committed_status"),
        ],
    )

    nr_rows = payloads["no_rfc"].get("rows") or []
    r = append_table(
        ws,
        r,
        f"3. No RFC (End FB <= Target FB; Rel Committed empty/not committed) — {len(nr_rows)} row(s)",
        [
            "Key",
            "Summary",
            "Status",
            "Competence Area",
            "Assignee",
            "End FB",
            "Target FB",
            "Rel Committed Status",
        ],
        nr_rows,
        [
            lambda rec: key_pair(rec),
            lambda rec: g(rec, "summary"),
            lambda rec: g(rec, "status"),
            lambda rec: g(rec, "competence_area"),
            lambda rec: g(rec, "assignee"),
            lambda rec: g(rec, "end_fb"),
            lambda rec: g(rec, "target_fb"),
            lambda rec: g(rec, "rel_committed_status"),
        ],
    )

    nc_rows = payloads["not_committed_2611"].get("rows") or []
    r = append_table(
        ws,
        r,
        f"4. Not committed at End FB 2611 (FB Committed Status open) — {len(nc_rows)} row(s)",
        [
            "Key",
            "Summary",
            "Competence Area",
            "Assignee",
            "End FB",
            "RC FB",
            "Risk Status",
            "Stretch Goal Reason",
        ],
        nc_rows,
        [
            lambda rec: key_pair(rec),
            lambda rec: g(rec, "summary"),
            lambda rec: g(rec, "competence_area"),
            lambda rec: g(rec, "assignee"),
            lambda rec: g(rec, "end_fb"),
            lambda rec: g(rec, "release_committed_fb"),
            lambda rec: g(rec, "risk_status"),
            lambda rec: g(rec, "stretch_goal_reason"),
        ],
    )

    autofit_columns(ws, 10)


def main() -> int:
    ap = argparse.ArgumentParser(description="Export J-Ant CB scopes to one multi-sheet xlsx.")
    ap.add_argument(
        "--workbook",
        type=Path,
        default=Path(r"C:/Users/kewang/OneDrive - Nokia/Projects/Tools/J-Ant_Latest - Copy.xlsm"),
        help="Source .xlsm path",
    )
    ap.add_argument(
        "--copy-to",
        type=Path,
        default=None,
        help="Optional local copy path if direct read hits PermissionError",
    )
    ap.add_argument(
        "--output-dir",
        type=Path,
        default=Path(r"C:\Users\kewang\OneDrive - Nokia\Projects\FOT\FOT report\DailyStatus"),
        help="Directory for the output xlsx",
    )
    ap.add_argument(
        "--features",
        nargs="+",
        default=["CB013987-SR", "CB014007", "14881", "15872"],
        help="One sheet per feature token (substring on Summary/Key)",
    )
    ap.add_argument(
        "--end-fb",
        default="2611",
        help="End FB week for not-committed filter",
    )
    ap.add_argument(
        "--file-prefix",
        default="J-Ant_CB_status",
        help="Output filename prefix before date",
    )
    args = ap.parse_args()

    analyzer = Path(__file__).resolve().parent / "analyze_j_ant_workbook.py"
    if not analyzer.is_file():
        sys.stderr.write(f"Missing analyzer script: {analyzer}\n")
        return 1

    args.output_dir.mkdir(parents=True, exist_ok=True)
    d = date.today().isoformat()
    out_path = args.output_dir / f"{args.file_prefix}_{d}.xlsx"

    copy_to = args.copy_to
    if copy_to is None:
        # Avoid cluttering the report folder; analyzer needs a writable copy if source is locked
        copy_to = Path(tempfile.gettempdir()) / f"J-Ant_read_temp_{d}.xlsm"

    used_names: set[str] = set()
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for feat in args.features:
        sheet_name = safe_sheet_name(feat, used_names)
        ws = wb.create_sheet(title=sheet_name)
        payloads = {
            "delayed": run_analyzer(
                Path(sys.executable), analyzer, args.workbook, copy_to, "delayed-rc-fb", feat, None
            ),
            "plan_deviation": run_analyzer(
                Path(sys.executable), analyzer, args.workbook, copy_to, "plan-deviation", feat, None
            ),
            "no_rfc": run_analyzer(
                Path(sys.executable),
                analyzer,
                args.workbook,
                copy_to,
                "no-rfc-end-before-target",
                feat,
                None,
            ),
            "not_committed_2611": run_analyzer(
                Path(sys.executable),
                analyzer,
                args.workbook,
                copy_to,
                "end-fb-not-committed",
                feat,
                args.end_fb,
            ),
        }
        build_sheet(ws, feat, payloads)

    wb.save(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
